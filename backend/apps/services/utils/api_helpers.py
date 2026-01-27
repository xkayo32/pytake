"""
API Response and Error Handling Utilities
"""
from rest_framework.response import Response
from rest_framework import status
from typing import Any, Dict, Optional


class APIResponse:
    """Standardized API response formatter"""

    @staticmethod
    def success(
        data: Any = None,
        message: str = "Operation successful",
        status_code: int = status.HTTP_200_OK,
        extra: Optional[Dict] = None,
    ) -> Response:
        """Return successful response"""
        response = {
            "success": True,
            "message": message,
            "data": data,
        }
        
        if extra:
            response.update(extra)
        
        return Response(response, status=status_code)

    @staticmethod
    def error(
        message: str,
        status_code: int = status.HTTP_400_BAD_REQUEST,
        errors: Optional[Dict] = None,
    ) -> Response:
        """Return error response"""
        response = {
            "success": False,
            "message": message,
        }
        
        if errors:
            response["errors"] = errors
        
        return Response(response, status=status_code)

    @staticmethod
    def paginated(
        data: list,
        page: int,
        page_size: int,
        total: int,
        message: str = "Data retrieved successfully",
    ) -> Response:
        """Return paginated response"""
        total_pages = (total + page_size - 1) // page_size
        
        return Response({
            "success": True,
            "message": message,
            "data": data,
            "pagination": {
                "page": page,
                "page_size": page_size,
                "total": total,
                "total_pages": total_pages,
                "has_next": page < total_pages,
                "has_previous": page > 1,
            },
        })


class APIError:
    """Custom API error classes"""

    class NotFound(Exception):
        """Resource not found"""
        def __init__(self, resource_type: str, resource_id: str = None):
            self.message = f"{resource_type} not found"
            if resource_id:
                self.message = f"{resource_type} {resource_id} not found"
            super().__init__(self.message)

    class Unauthorized(Exception):
        """Unauthorized access"""
        def __init__(self, message: str = "Unauthorized"):
            self.message = message
            super().__init__(message)

    class Forbidden(Exception):
        """Access forbidden"""
        def __init__(self, message: str = "Access forbidden"):
            self.message = message
            super().__init__(message)

    class ValidationError(Exception):
        """Validation error"""
        def __init__(self, message: str, errors: Dict = None):
            self.message = message
            self.errors = errors or {}
            super().__init__(message)

    class IntegrationError(Exception):
        """External integration error"""
        def __init__(self, provider: str, message: str):
            self.message = f"{provider} integration error: {message}"
            self.provider = provider
            super().__init__(self.message)


class SearchHelper:
    """Search and filter utilities"""

    @staticmethod
    def build_search_query(search_term: str, fields: list):
        """Build Q object for multi-field search"""
        from django.db.models import Q
        
        if not search_term:
            return Q()
        
        query = Q()
        for field in fields:
            query |= Q(**{f"{field}__icontains": search_term})
        
        return query

    @staticmethod
    def apply_filters(queryset, filters: dict, filter_fields: dict):
        """Apply multiple filters to queryset"""
        for field, value in filters.items():
            if field in filter_fields and value:
                queryset = queryset.filter(**{field: value})
        
        return queryset

    @staticmethod
    def apply_ordering(queryset, order_by: str, allowed_fields: list):
        """Apply ordering safely"""
        if order_by and order_by.lstrip('-') in allowed_fields:
            return queryset.order_by(order_by)
        
        return queryset


class PaginationHelper:
    """Pagination utilities"""

    DEFAULT_PAGE_SIZE = 50
    MAX_PAGE_SIZE = 100

    @staticmethod
    def paginate(queryset, page: int = 1, page_size: int = None):
        """Paginate queryset"""
        if page_size is None:
            page_size = PaginationHelper.DEFAULT_PAGE_SIZE
        
        page_size = min(page_size, PaginationHelper.MAX_PAGE_SIZE)
        
        if page < 1:
            page = 1
        
        offset = (page - 1) * page_size
        
        return {
            'items': queryset[offset:offset + page_size],
            'page': page,
            'page_size': page_size,
            'total': queryset.count(),
            'offset': offset,
        }


class ExcelExporter:
    """Export data to Excel"""

    @staticmethod
    def export_contacts(contacts: list, filename: str = "contacts.xlsx"):
        """Export contacts to Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Contacts"
            
            # Headers
            headers = ["Name", "Email", "Phone", "WhatsApp ID", "Lifecycle", "Created At"]
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
            
            # Data
            for row, contact in enumerate(contacts, 2):
                ws.cell(row=row, column=1).value = contact.name
                ws.cell(row=row, column=2).value = contact.email
                ws.cell(row=row, column=3).value = contact.phone
                ws.cell(row=row, column=4).value = contact.whatsapp_id
                ws.cell(row=row, column=5).value = contact.lifecycle_stage
                ws.cell(row=row, column=6).value = contact.created_at
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(filename)
            return filename
        except ImportError:
            raise ImportError("openpyxl required for Excel export")


class CSVExporter:
    """Export data to CSV"""

    @staticmethod
    def export_conversations(conversations: list, filename: str = "conversations.csv"):
        """Export conversations to CSV"""
        import csv
        
        with open(filename, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            
            # Headers
            writer.writerow([
                "ID",
                "Contact",
                "Status",
                "Messages",
                "Created At",
                "Updated At",
            ])
            
            # Data
            for conv in conversations:
                writer.writerow([
                    str(conv.id),
                    conv.contact.name if conv.contact else "Unknown",
                    conv.status,
                    conv.message_set.count(),
                    conv.created_at.isoformat(),
                    conv.updated_at.isoformat(),
                ])
        
        return filename
